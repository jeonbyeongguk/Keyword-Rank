import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import base64

# 순위 선정 함수
def search_and_fill_excel(file_path,keyword):
    # 엑셀 파일 불러오기
    df = pd.read_excel(file_path)

    # 2열에서 "keyword" 찾기
    found_cats_col2 = df.iloc[:, 1].str.contains(keyword, na=False)

    # 3열에서 "keyword" 찾기
    found_cats_col3 = df.iloc[:, 2].str.contains(keyword, na=False)

    # 4열에서 "keyword" 찾기
    found_cats_col4 = df.iloc[:, 3].str.contains(keyword, na=False)

    # 5열에 값을 할당
    df.loc[found_cats_col2, "순위"] = 1
    df.loc[found_cats_col3, "순위"] = 2
    df.loc[found_cats_col4, "순위"] = 3
    df.loc[~(found_cats_col2 | found_cats_col3 | found_cats_col4), "순위"] = float("nan")  # "없음"을 실수형으로 변환

    # 변경된 데이터프레임을 엑셀 파일로 저장
    df.to_excel(file_path, index=False)

# 다운로드 함수
def get_table_download_link(df, input_keyword, file_name=None):
    if file_name is None:
        file_name = f'{input_keyword} PC 순위 결과.xlsx'
    
    with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
        df.to_excel(excel_writer, index=False)
    
    with open(file_name, 'rb') as file:
        data = file.read()
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}"> 엑셀 파일 다운로드(XLSX)</a>'
    return href


def search_keyword(keyword):
    url = f'https://search.naver.com/search.naver?query={keyword}'
    response = requests.get(url)
    
    if response.status_code == 200:
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        title1 = soup.select_one('#power_link_body > ul > li:nth-child(1) > div > div.title_url_area > a > span')
        title2 = soup.select_one('#power_link_body > ul > li:nth-child(2) > div > div.title_url_area > a > span')
        title3 = soup.select_one('#power_link_body > ul > li:nth-child(3) > div > div.title_url_area > a > span')

        # Check if any of the titles are None before calling get_text()
        title1_text = title1.get_text() if title1 else None
        title2_text = title2.get_text() if title2 else None
        title3_text = title3.get_text() if title3 else None

        print(f"{row[0].row -1}번" f"{keyword}: {title1_text}, {title2_text}, {title3_text}")

        return title1_text, title2_text, title3_text

    elif response.status_code == 403:
        st.write("요청이 많아 5분 후 다시 시작합니다.")
        time.sleep(240)  # Wait for 1 minute
        return search_keyword(keyword)  # Retry the search

    else:
        st.write(f"Error occurred with status code: {response.status_code}")
        return None, None, None


# 페이지 기본 설정
st.set_page_config(
    page_icon="💡",
    page_title="[PC]키워드 순위 찾기 베타",
    layout="wide",
)

st.title("💡[PC]키워드 순위 찾기 베타버전")

uploaded_file = st.file_uploader("'CSV' 또는 'xlsx' 파일 업로드하세요!!", type=['csv', 'xlsx'])

if uploaded_file is not None:
    try:
        # 업로드된 파일을 데이터프레임으로 읽기
        if uploaded_file.name.endswith('csv'):
            df = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith(('xls', 'xlsx')):
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        else:
            st.error('올바른 파일 형식이 아닙니다. CSV 또는 엑셀 파일을 업로드해주세요.')
            st.stop()

        # 데이터프레임 표시
        st.write(df)

        input_keyword = st.text_input('검색할 키워드를 입력해주세요:', key='input_keyword')

        if input_keyword:
            st.success(f"입력된 키워드:'{input_keyword}'")
        
            # 실행하기 버튼 추가
            if st.button("실행하기"):
                
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

                # Open Excel file
                file_path = uploaded_file
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                start_time = time.time()

                iteration_count = 0

                # Read keywords from Excel file and search
                for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=False):
                    iteration_count += 1
                    keyword = row[0].value
                    title1_result, title2_result, title3_result = search_keyword(keyword)
                    sheet.cell(row=row[0].row, column=2, value=title1_result)
                    sheet.cell(row=row[0].row, column=3, value=title2_result)
                    sheet.cell(row=row[0].row, column=4, value=title3_result)
                    workbook.save(file_path)
                    st.write( f"{row[0].row -1}번", f"검색키워드: {keyword}, 제목: {title1_result}, {title2_result}, {title3_result}")

                # Save the updated Excel file
                workbook.save(file_path)
                workbook.close()

                end_time = time.time()
                execution_time = round(end_time - start_time,1)

                st.write("찾은 키워드 개수:", iteration_count)
                st.write("총 실행 시간:", execution_time, "초")
                
                # search_keyword 함수가 끝난 후에 search_and_fill_excel 함수 실행
                search_and_fill_excel(file_path, input_keyword)
                
                # 데이터프레임 다시 불러오기
                df = pd.read_excel(file_path)
                
                # 데이터프레임 표시
                st.write(df)
                
                # 데이터프레임 다운로드 버튼
                st.markdown(get_table_download_link(df, input_keyword), unsafe_allow_html=True)

    except Exception as e:
        st.error(f'오류 발생: {e}')
