import requests
from bs4 import BeautifulSoup
import openpyxl
import time

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}


def search_keyword(keyword):
    url = f'https://m.search.naver.com/search.naver?query={keyword}'
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        title1 = soup.select_one('#power_link_body > li:nth-child(1) > div > div.tit_wrap > div > a > div.tit_area > span:nth-child(1) > mark')
        title2 = soup.select_one('#power_link_body > li:nth-child(2) > div > div.tit_wrap > div > a > div.tit_area > span:nth-child(1) > mark')
        title3 = soup.select_one('#power_link_body > li:nth-child(3) > div > div.tit_wrap > div > a > div.tit_area > span:nth-child(1) > mark')

        # Check if any of the titles are None before calling get_text()
        title1_text = title1.get_text() if title1 else None
        title2_text = title2.get_text() if title2 else None
        title3_text = title3.get_text() if title3 else None

        print(f"{row[0].row -1}번" f"{keyword}: {title1_text}, {title2_text}, {title3_text}")

        return title1_text, title2_text, title3_text

    elif response.status_code == 403:
        print("403 Forbidden Error occurred. Waiting for 5 minute before retrying.")
        time.sleep(240)  # Wait for 1 minute
        return search_keyword(keyword)  # Retry the search

    else:
        print(f"Error occurred with status code: {response.status_code}")
        return None, None, None

# Open Excel file
file_path = r'C:\Users\KID_전병국\Desktop\신입파일\키워드 비딩 입력 자료.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

start_time = time.time()

iteration_count = 0

# Read keywords from Excel file and search
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=False):
    iteration_count += 1
    keyword = row[0].value
    title1_result, title2_result, title3_result = search_keyword(keyword)
    sheet.cell(row=row[0].row, column=2, value=title1_result)
    sheet.cell(row=row[0].row, column=3, value=title2_result)
    sheet.cell(row=row[0].row, column=4, value=title3_result)
    workbook.save(file_path)
    

# Save the updated Excel file
workbook.save(file_path)
workbook.close()

end_time = time.time()
execution_time = end_time - start_time

print("Total iterations:", iteration_count)
print("코드 실행 시간:", execution_time, "초")
print("완료됐습니다.")
